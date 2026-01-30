from pathlib import Path
import ast
from openpyxl.reader.excel import load_workbook


# 封装的读取表格文件方法
def read_test_data_from_excel(
        file_path,
        sheet_name,
        parse_sku,  # 开关：是否解析SKU列
        sku_col_index  # SKU列的索引
):
    """
    通用Excel测试数据读取函数（底层配置）
    :param file_path: Excel文件路径（绝对/相对）
    :param sheet_name: 工作表名称
    :param parse_sku: 是否解析SKU列（True=解析多SKU，False=不解析）
    :param sku_col_index: SKU列的索引
    :return: 解析后的测试数据列表
    """
    # 处理路径：兼容相对/绝对路径
    file_path = Path(file_path).absolute()
    if not file_path.exists():
        raise FileNotFoundError(f"Excel文件不存在：{file_path}")

    workbook = load_workbook(filename=file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表{sheet_name}不存在，可选工作表：{workbook.sheetnames}")

    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过标题行（第1行）
        row_parsed = list(row)  # 转为列表，方便修改

        # 仅当需要解析SKU时，执行SKU解析逻辑
        if parse_sku:
            sku_raw = row_parsed[sku_col_index] if len(row_parsed) > sku_col_index else None
            if sku_raw:
                # 格式1：列表字符串（如["SKU001","SKU002"]）→ 解析为列表
                if isinstance(sku_raw, str) and sku_raw.startswith("[") and sku_raw.endswith("]"):
                    sku_list = ast.literal_eval(sku_raw)
                # 格式2：逗号分隔字符串（如SKU001,SKU002）→ 分割为列表
                elif isinstance(sku_raw, str) and "," in sku_raw:
                    sku_list = [s.strip() for s in sku_raw.split(",")]
                # 格式3：单个SKU（如SKU001）→ 转为单元素列表
                else:
                    sku_list = [sku_raw]
            else:
                sku_list = []  # 空值处理
            # 替换SKU列为解析后的列表
            row_parsed[sku_col_index] = sku_list

        # 转回元组（保持不可变，符合原有逻辑）
        data.append(tuple(row_parsed))

    workbook.close()  # 关闭工作簿，释放资源
    return data