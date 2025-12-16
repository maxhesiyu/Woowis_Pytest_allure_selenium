from pathlib import Path
from time import sleep

import allure
import pytest
from openpyxl.reader.excel import load_workbook
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import wait

from common.base import sel_end_keys, sel_click, assert_text_in_element, refresh_when_element_appears
from common.log import log
from config.config import ENV

# 新增：获取项目根目录（适配任意执行路径）
PROJECT_ROOT = Path(__file__).parent.parent.absolute()
# 新增：Excel文件绝对路径
EXCEL_FILE_PATH = PROJECT_ROOT / "测试登录参数化.xlsx"

# ========== 读取Excel数据 ==========
def read_test_data_from_excel(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行
        data.append(row)
    return data

# ========== 测试类 ==========
class TestLongin:
    @pytest.mark.parametrize(
        "test_case",
        read_test_data_from_excel(str(EXCEL_FILE_PATH),
                                  'Sheet1'),
        ids=('login01', 'login02', 'login03')
    )
    @allure.feature('登录注册')
    @allure.story('登录测试用例')
    def test_login(self, test_case, open_page):
        """修复：简化逻辑，确保能执行"""
        with allure.step("获取测试数据"):
            case_name = test_case[0]
            pcno = test_case[1]
            password = test_case[2]
            expected_result = test_case[3]
        driver = open_page
        allure.dynamic.title(f"登录测试：{case_name}")
        print(f"📌 执行用例：{case_name}，账号：{pcno}，密码：{password}")
        try:
            with allure.step(f"输入顾客编号: {pcno}"):
                sel_end_keys(driver, (By.XPATH, "//input[@placeholder='顾客编号(PC ID)']"), pcno)
            with allure.step(f"输入密码: {password}"):
                sel_end_keys(driver, (By.XPATH, "//input[@placeholder='密码(Password)']"), password)
            # ========== 分场景断言 ==========
            # 场景1：顾客编号格式不正确
            if '顾客编号格式不正确(Invalid PC ID)' in expected_result:
                with allure.step(f"断言：提示文本包含「{expected_result}」"):
                    with allure.step("点击登录"):
                        sel_click(driver, (By.XPATH, "//span[contains(text(),'登录(Login)')]"))
                    assert_text_in_element(
                        driver=driver,
                        locator=(By.XPATH,
                                     "//div[@class='ivu-form-item ivu-form-item-required ivu-form-item-error']//div[2]"),
                        target_text=expected_result,
                    )
                log.info(f"✅ 用例「{case_name}」断言通过（顾客编号格式错误场景）")

            # 场景2：登录成功（显示指定用户信息）
            elif '何思宇 (60003152)' in expected_result:
                with allure.step(f"断言：登录后显示用户信息「{expected_result}」"):
                    with allure.step("点击登录"):
                        sel_click(driver, (By.XPATH, "//span[contains(text(),'登录(Login)')]"))
                    assert_text_in_element(
                        driver=driver,
                        locator=(By.XPATH, "//span[contains(text(),'何思宇 (60003152)')]"),
                        target_text=expected_result,
                    )
                # 登录成功后执行退出操作，还原测试环境
                with allure.step("执行退出登录，清理测试环境"):
                    sel_click(driver,(By.XPATH,"//span[contains(text(),'退出登录(Log out)')]"))
                log.info(f"✅ 用例「{case_name}」断言通过（登录成功场景）")

            # 场景3：密码不能为空
            elif '密码不能为空 (Password is missing)' in expected_result:
                with allure.step(f"断言：密码为空提示包含「{expected_result}」"):
                    driver.find_element(By.XPATH, "//input[@placeholder='密码(Password)']").send_keys(Keys.CONTROL, 'a')
                    driver.find_element(By.XPATH, "//input[@placeholder='密码(Password)']").send_keys(Keys.DELETE)
                sleep(1)
                sel_click(driver, (By.XPATH, "//span[contains(text(),'登录(Login)')]"))
                assert_text_in_element(
                        driver=driver,
                        locator=(By.XPATH, "//div[@class='ivu-form-item-error-tip']"),
                        target_text=expected_result,
                )
                log.info(f"✅ 用例「{case_name}」断言通过（密码为空场景）")

            # 场景4：其他未匹配的预期结果（兜底提示）
            else:
                raise ValueError(f"未匹配到对应的断言场景，预期结果：{expected_result}")
        except Exception as e:
        # 断言失败/执行异常时，捕获并附加截图到Allure报告
            allure.attach(
                driver.get_screenshot_as_png(),
                f"用例执行失败：{case_name}",
                allure.attachment_type.PNG
            )
            log.error(f"❌ 用例「{case_name}」执行失败：{str(e)}")
            raise  # 重新抛出异常，让pytest标记用例失败


    @allure.story('用户登录')
    def test_shopping_01(self, DengLu):
        """修复：简化断言，确保能执行"""
        with allure.step("断言结果故意错误"):
            driver = DengLu
            allure.dynamic.title("登录后用户名验证")
            assert_text_in_element(driver,
                                   (By.XPATH, f"//span[contains(text(),'何思宇 ({ENV.pcno})')]"),"测试")

    @allure.story('用户注册流程')
    def test_ZhuCe_01(self,open_page):
        with allure.step("注册PC用户"):
            driver = open_page
            allure.dynamic.title("注册基本流程到个人信息填写")
            sel_click(driver,(By.XPATH,"//span[contains(text(),'点击注册(Click To Register)')]"))
            sel_click(driver,(By.XPATH,"//div[contains(text(),'申请优惠顾客')]"))
            allure.dynamic.title("弹窗同意注册提示")
            # 弹窗同意注册提示
            sel_click(driver, (By.XPATH, "//span[contains(text(),'我理解并同意以下全部内容')]"))
            sel_click(driver,(By.XPATH,"//button[@class='button confirm ivu-btn ivu-btn-primary']"))
            allure.dynamic.title("手机号填写")
            sel_end_keys(driver, (By.XPATH, "//body/div[@id='app']/div[@class='root-content']/div[@id='root']/div[@class='container createAccount']/div[@class='content']/section/form[@class='ivu-form ivu-form-label-top']/div[@class='c-form-item ivu-form-item ivu-form-item-required']/div[@class='ivu-form-item-content']/div[@class='ivu-input-wrapper ivu-input-wrapper-large ivu-input-type']/input[1]"),ENV.randomPhone)
            allure.dynamic.title("验证码填写")
            sel_end_keys(driver, (By.XPATH, "//div[@class='yzmInput ivu-col ivu-col-span-11']//input[@type='text']"),ENV.CAPTCHA)
            allure.dynamic.title("朋友号填写")
            sel_click(driver, (By.XPATH,"//label[contains(text(),'朋友(Friend)')]"))
            sleep(1)
            sel_end_keys(driver, (By.XPATH, "//input[@placeholder='顾客编号(PC ID)']"),ENV.referrer)
            sleep(1)
            allure.dynamic.title("验证朋友号")
            sel_click(driver, (By.XPATH, "//span[contains(text(),'验证顾客编号(Verify the PC ID)')]"))
            sel_click(driver, (By.XPATH, "//div[contains(@class,'bottom-button button1')]//button[2]"))
            allure.dynamic.title("个人信息填写")
            sleep(1)
            allure.dynamic.title("姓名填写")
            sel_end_keys(driver, (By.XPATH, "//input[@placeholder='请填写真实姓名 (Please fill in the legal name)']"), ENV.name)
            allure.dynamic.title("密码填写")
            sel_end_keys(driver, (By.XPATH, "//div[@class='passwordlabel ivu-form-item ivu-form-item-required']//input[@type='password']"), ENV.npwd)
            sel_end_keys(driver, (By.XPATH, "//div[@class='ivu-form-item ivu-form-item-required']//input[@type='password']"),ENV.ncpwd)
            allure.dynamic.title("国外身份证填写")
            sleep(2)
            sel_click(driver, (By.XPATH, "//label[contains(text(),'其它国籍（地区）身份证件')]"))
            sel_end_keys(driver, (By.XPATH, "//div[@class='ivu-form-item']//div[@class='ivu-input-wrapper ivu-input-wrapper-large ivu-input-type']//input[@type='text']"),ENV.randomSFZ)
            allure.dynamic.title("日历选择日期")
            # 直接通过JS查找所有日期输入框，设置值（适配任意定位）
            driver.execute_script("""
                // 查找包含生日/生日相关的input（按placeholder/name匹配）
                const birthdayInputs = document.querySelectorAll('input[placeholder*="生日"], input[name*="birthday"], .selectbirthday input');
                if (birthdayInputs.length > 0) {
                    birthdayInputs[0].removeAttribute('readonly');
                    birthdayInputs[0].value = '2005-06-17';
                    birthdayInputs[0].dispatchEvent(new Event('change'));
                }
            """)
            allure.dynamic.title("勾选优惠顾客政策并点击申请优惠注册按钮")
            sleep(1)
            sel_click(driver, (By.XPATH, "//a[contains(text(),'优惠顾客政策(Preferential customer policy)')]"))
            sleep(1)
            sel_click(driver, (By.XPATH, "//button[@class='goon-btn ivu-btn ivu-btn-primary ivu-btn-large']"))
            sel_click(driver, (By.XPATH, "//button[@class='gobtn ivu-btn ivu-btn-primary ivu-btn-large']"))
            sel_click(driver, (By.XPATH, "//span[contains(text(),'开始购物')]"))
            refresh_when_element_appears(driver,(By.XPATH, "//i[@class='ivu-icon ivu-icon-ios-close-circle-outline']"),
                                         (By.XPATH, "//input[@placeholder='搜索(Search)']"))










