import ast
from pathlib import Path
from time import sleep
import pytest
from openpyxl.reader.excel import load_workbook
import allure
from selenium.webdriver.common.by import By

from common.base import sel_end_keys, sel_click, refresh_when_element_appears, redirect_URL, \
    click_element_if_exists_with_wait, sel_hover, assert_text_in_element, check_text_exists
from po.shopping import Shopping_product_Order, Shopping_querySku

# 新增：获取项目根目录（适配任意执行路径）
PROJECT_ROOT = Path(__file__).parent.parent.absolute()
# 新增：Excel文件绝对路径
EXCEL_FILE_PATH = PROJECT_ROOT / "促销赠品获取参数化.xlsx"

# ========== 读取Excel数据 ==========
def read_test_data_from_excel(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 第一行是标题行
        # 新增：解析多SKU（兼容2种格式：列表字符串/逗号分隔字符串）
        sku_raw = row[4]  # SKU列
        if sku_raw:
            # 格式1：列表字符串（如["SKU001","SKU002"]）→ 解析为列表
            if isinstance(sku_raw, str) and sku_raw.startswith("[") and sku_raw.endswith("]"):
                sku_list = ast.literal_eval(sku_raw)
            # 格式2：逗号分隔字符串（如SKU001,SKU002）→ 分割为列表
            elif isinstance(sku_raw, str) and "," in sku_raw:
                sku_list = [s.strip() for s in sku_raw.split(",")]
            # 格式3：单个SKU（如SKU001）→ 转为单元素列表
            else:
                sku_list = [sku_raw]
        else:
            sku_list = []  # 空值处理

        # 替换原SKU字段为解析后的列表，其他数据保留
        row_parsed = list(row)
        row_parsed[4] = sku_list  # 第5列（索引4）改为SKU列表
        data.append(tuple(row_parsed))
    return data



@allure.title('用户获取不同赠品')
class TestFreeGift:

    @pytest.mark.parametrize(
        "test_case",
        read_test_data_from_excel(str(EXCEL_FILE_PATH),
                                  '促销')
    )
    @allure.title('用户获取不同促销')
    def test_shopping_FreeGift(self,test_case, open_page):
        with allure.step('获取测试数据'):
            case_name = test_case[0]  # 场景名称
            pcno = test_case[1]  # 账号
            password = test_case[2]  # 密码
            skuTime = test_case[3]  # 需要循环购买sku的次数
            sku_list = test_case[4]  # 解析后的SKU列表（单个/多个）
            expected_result = test_case[5] #预期结果
            # actual_result = test_case[4] #实际结果
            driver = open_page
        allure.dynamic.title(f"测试：{case_name}")
        with allure.step(f"输入顾客编号: {pcno}"):
            sel_end_keys(driver, (By.XPATH, "//input[@placeholder='顾客编号(PC ID)']"), pcno)
        with allure.step(f"输入密码: {password}"):
            sel_end_keys(driver, (By.XPATH, "//input[@placeholder='密码(Password)']"), password)
        with allure.step(f"{pcno}登录"):
            sel_click(driver, (By.XPATH, "//span[contains(text(),'登录(Login)')]"))
        # with allure.step("防止隐私协议弹窗阻碍流程"):
        #     click_element_if_exists_with_wait(driver, (By.XPATH, "//span[contains(text(),'我理解并同意以下全部内容')]"))
        #     click_element_if_exists_with_wait(driver, (By.XPATH, "//button[@class='ivu-btn ivu-btn-primary']//span[contains(text(),'同意')]"))
        with allure.step("用户下单流程"):
            allure.dynamic.title("用户下单流程")
            allure.step("重定向页面")
            # 防止点数弹窗拦截把点击订购的按钮拦截掉
            sleep(1)
            refresh_when_element_appears(driver, (By.XPATH, "//span[contains(text(),'确定')]"),
                                         (By.XPATH, "//span[@class='main zh'][contains(text(),'在线订购')]"))
            # 重定向URL
            redirect_URL(driver, "order/product",(By.XPATH, "//span[@class='main zh'][contains(text(),'在线订购')]"))
            with allure.step("点击产品选项,进行产品加入购物车"):
                sel_click(driver, (By.XPATH, "//a[@class='top-link']//span[@class='zh'][contains(text(),'产品')]"))
            with allure.step("加购产品循环次数"):
                for i in range(skuTime):
                    Shopping_querySku(driver, sku_list)
            with allure.step("鼠标悬停在购物车元素上（不点击）"):
                sleep(1)
                sel_hover(driver, (By.XPATH, "//a[@class='item']//span[@class='zh'][contains(text(),'购物车')]"))
                sleep(1)  # 防止结算按钮出不来
            with allure.step("点击购物车的去结算按钮）"):
                sel_click(driver, (By.XPATH, "//span[contains(text(),'去结算(Go to pay)')]"))
                allure.step("全局查找预期结果")
                # if check_text_exists(driver, expected_result):
                #     print("文本存在于页面当中")
                # else:
                #     print("需要查找的文字内容不存在于页面中")
                # 1. 检查预期文本是否存在
                text_exists = check_text_exists(driver, expected_result)
                if text_exists:
                    # 文本存在 → 用例通过，记录日志到Allure
                    success_msg = f"✅ 用例[{case_name}]成功：预期文本「{expected_result}」存在于页面中"
                    print(success_msg)
                    allure.attach(success_msg, "结果验证", allure.attachment_type.TEXT)
                else:
                    # 文本不存在 → 标记用例失败（触发钩子自动截图）
                    fail_msg = f"❌ 用例[{case_name}]失败：预期文本「{expected_result}」不存在于页面中"
                    print(fail_msg)
                    # 可选：额外添加文本日志到Allure（截图由钩子自动完成）
                    allure.attach(fail_msg, "结果验证", allure.attachment_type.TEXT)
                    # 关键：调用pytest.fail标记用例失败，钩子会自动捕获失败状态并截图
                    pytest.fail(fail_msg)
                #
                # allure.step("查找文字后断言是否符合预期")
                # assert_text_in_element(driver=driver,
                #                        locator=(By.XPATH, f"//span[contains(text(),'{expected_result}')]"),
                #                        target_text=expected_result)











